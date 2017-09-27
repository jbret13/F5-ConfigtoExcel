from openpyxl import Workbook, load_workbook

fname = input('Enter input file name: ')
dname = input('Enter output file name: ')

try:
    in_file = open(fname)
except:
    print('File cannot be opened:', fname)
    quit()

try:
    open(dname)
except:
    wb = Workbook()
    wb.save(dname)

wb = load_workbook(dname)

#Create and name all worksheets 
ws1 = wb.create_sheet('Virtual Servers', 0)
ws2 = wb.create_sheet('Pools', 1)
ws3 = wb.create_sheet('Members', 2)
ws4 = wb.create_sheet('Nodes', 3)
ws5 = wb.create_sheet('Monitors', 4)
ws6 = wb.create_sheet('Profiles', 5)
ws7 = wb.create_sheet('Persistence', 6)
ws8 = wb.create_sheet('SNAT Pool', 7)
ws9 = wb.create_sheet('iRule', 8)
ws10 = wb.create_sheet('DataGroup', 9)

#Create headers for the worksheets
vs_headers = {'A' : 'Virtual Server Name', 'B' : 'IP Address', 'C' : 'Port', 'D': 'Destination',
            'E' : 'Mask', 'F' : 'Pool', 'G' : 'SNAT Pool', 'H' : 'VLAN', 'I' : 'IP Protocol', 
            'J' : 'Persistence', 'K' : 'iRule', 'L' : 'Client Profile', 'M' : 'HTTP Profile', 
            'N' : 'One Connect Profile', 'O' : 'NTLM Profile', 'P' : 'Client SSL Profile',
            'Q' : 'Server SSL Profile'}
pool_headers = { 'A' : 'Pool Name', 'B' : 'Load Balancing', 'C' : 'Monitor 1', 'D' : 'Monitor 2',
              'E' : 'Priority Group', 'F' : 'Member 1', 'G' : 'Priority M1', 'H' : 'Member 2',
              'I' : 'Priority M2', 'J' : 'Member 3', 'K' : 'Member 4', 'L' : 'Member 5', 'M' : 'Member 6', 
              'N' : 'Member 7', 'O' : 'Member 8', 'P' : 'Member 9', 'Q' : 'Member 10', 'R' : 'Member 11', 
              'S' : 'Member 12'}
member_headers = {'A' : 'Member', 'B' : 'IP Address', 'C' : 'Port'}
node_headers = {'A' : 'Node', 'B' : 'Description', 'C' : 'Monitor 1', 'D' : 'Monitor 2'}
monitor_headers = {'A' : 'Monitor', 'B' : 'Default From', 'C' : 'Send String', 'D' : 'Receive String'}
profile_headers = {'A' : 'Profile', 'B' : 'Type', 'C' : 'Defaults From', 'D' : 'Options'}
persistence_headers = {'A' : 'Profile', 'B' : 'Type', 'C' : 'Defaults From', 'D' : 'Options'}
snatpool_headers = {'A' : 'Name', 'B' : 'Member', 'C' : 'Member'}
irule_headers = {'A' : 'Name', 'B' : 'Partition', 'C' : ' Data'}
datagroup_headers = {'A' : 'Name', 'B' : 'Partition', 'C' : 'Type', 'D' : 'Members'}

#Add the headers to the worksheet
ws1.append(vs_headers)
ws2.append(pool_headers)
ws3.append(member_headers)
ws4.append(node_headers)
ws5.append(monitor_headers)
ws6.append(profile_headers)
ws7.append(persistence_headers)
ws8.append(snatpool_headers)
ws9.append(irule_headers)
ws10.append(datagroup_headers)

cell = 2
pcolumn = 1
prscolumn = 1


for line in in_file:
    if line.startswith('virtual') and 'address' not in line:
        vsname = line.split()[1].strip()
        vs = {}
        while line.startswith('}') is False:
            vs['A'] = vsname
            if 'snatpool' in line:
                snatpool = line.split()[-1].strip()
                vs['G'] = snatpool
            elif 'pool ' in line:
                pool = line.split()[-1].strip()
                vs['F'] = pool
            elif ' destination' in line:
                ipaddr = line.split(':')[0].split()[1].strip()
                port = line.split(':')[1].strip()
                if port == 'any':
                    port = '0'
                elif port == 'https':
                    port = '443'
                elif port == 'ddm-rdb':
                    port = '446'
                elif port == 'http':
                    port = '80'
                elif port == 'pharos':
                    port = '4443'
                elif port == 'saris':
                    port = '4442'
                vs['B'] = ipaddr
                vs['C'] = port
                dest = ipaddr + ':' + port
                vs['D'] = dest
            elif 'rules ' in line:
                irule = line.split()[-1].strip()
                vs['K'] = irule
            elif 'persist ' in line:
                persist = line.split()[-1].strip()
                vs['J'] = persist
            elif 'vlans ' in line:
                vlans = line.split()[1].strip()
                vs['H'] = vlans
            elif 'ip protocol' in line:
                protocol = line.split()[-1].strip()
                vs['I'] = protocol
            elif 'profiles' in line:
                vsprflcol = 'L'
                start_bracket, end_bracket = 1, 2
                while start_bracket != end_bracket:
                    if start_bracket == 1 and end_bracket == 2:
                        start_bracket = 0
                        end_bracket = 0
                    if 'profiles' in line and '}' in line:
                        vs[vsprflcol] = line.split()[1].strip()
                    if '{' in line and 'profiles' not in line:
                        profile = line.split()[0].strip()
                        if 'tcp' in line:
                            vs['L'] = profile
                            vsprflcol = chr(ord(vsprflcol)+1)
                        elif 'http' in line or ' test ' in line:
                            vs['M'] = profile
                            vsprflcol = chr(ord(vsprflcol)+1)
                        elif 'oneconnect' in line:
                            vs['N'] = profile
                        elif 'ntlm' in line:
                            vs['O'] = profile
                        elif 'client' in line:
                            vs['P'] = profile
                        elif 'server' in line:
                            vs['Q'] = profile
                        else:
                            vs['S'] = line.split()[0].strip()
                    vsprflcol = chr(ord(vsprflcol)+1)
                    if '{' in line:
                        for i in line:
                            if i == '{':
                                start_bracket += 1
                    if '}' in line:
                        for i in line:
                            if i == '}':
                                end_bracket += 1
                    if start_bracket != end_bracket:
                        line = next(in_file)
            line = next(in_file)
        ws1.append(vs)
    if line.startswith('pool '):
        poolname = line.split()[1].strip()
        pool = {}
        pool['A'] = poolname
        lbmethod = ''
        while line.startswith('}') is False:
            if 'lb method member least conn' in line:
                lbmethod = 'least-connections-member'
            elif 'lb method least conn' in line:
                lbmethod = 'least-connections-node'
            elif 'lb method member observed' in line:
                lbmethod = 'observed-member'
            elif 'lb method observed' in line:
                lbmethod = 'observed-node'
            elif 'lb method member ratio' in line:
                lbmethod = 'ratio-least-connections-member'
            if 'monitor all' in line and 'and' not in line:
                #if line.split()[-1].strip() == 'gateway_icmp' or line.split()[-1].strip() == 'http' or line.split()[-1].strip() == 'https'or line.split()[-1].strip() == 'https_443':
                #    pool['C'] = '/Common/' + line.split()[-1].strip()
                #else:
                pool['C'] =  line.split()[-1].strip()
            elif 'monitor all' and 'and' in line:
                #if line.split()[2].strip() == 'gateway_icmp' or line.split()[2].strip() == 'http' or line.split()[-1].strip() == 'https' or line.split()[-1].strip() == 'https_443':
                    #pool['C'] = '/Common/' + line.split()[2].strip()
                #else:
                pool['C'] = line.split()[2].strip()
                #if line.split()[-1].strip() == 'gateway_icmp' or line.split()[-1].strip() == 'http' or line.split()[-1].strip() == 'https' or line.split()[-1].strip() == 'https_443':
                    #pool['D'] = '/Common/' + line.split()[-1].strip()
                #else:
                pool['D'] =  line.split()[-1].strip()
            if 'min active members' in line:
                pool['E'] = line.lstrip().rstrip()
            if 'members {' in line:
                start_bracket, end_bracket = 1, 2
                members, port, exlcol = {}, {}, 'F'
                while start_bracket != end_bracket:
                    if start_bracket == 1 and end_bracket == 2:
                        start_bracket = 0
                        end_bracket = 0
                    if ':' in line:
                        address = line.split(':')[0].split()[-1].strip()
                        port = line.split(':')[1].split()[0].strip()
                        if port == 'any':
                            port = '0'
                        elif port == 'https':
                            port = '443'
                        elif port == 'ddm-rdb':
                            port = '446'
                        elif port == 'http':
                            port = '80'
                        elif port == 'pharos':
                            port = '4443'
                        elif port == 'saris':
                            port = '4442'
                        colon = ':'
                        member = address + colon + port
                        members[exlcol] = member
                        if exlcol == 'F' or exlcol == 'H':
                            exlcol = chr(ord(exlcol)+2)
                        elif exlcol != 'Z' and exlcol != 'Y' and len(exlcol) == 1:
                            exlcol = chr(ord(exlcol)+1)
                            #print(exlcol)
                        elif exlcol == 'Z' or exlcol == 'Y':
                            #print('Should be AA ' + exlcol)
                            exlcol = 'AA'
                        elif len(exlcol) == 2:
                            exlcol = 'A'+ chr(ord(exlcol[1]) + 1)
                    if 'priority' in line:
                        exlcol = chr(ord(exlcol)-1)
                        members[exlcol] = line.lstrip().rstrip()
                        exlcol = chr(ord(exlcol)+1)
                    if '{' in line:
                        for i in line:
                            if i == '{':
                                start_bracket += 1
                                #print('start_bracket', start_bracket)
                    if '}' in line:
                        for i in line:
                            if i == '}':
                                end_bracket += 1
                                #print('end_bracket', end_bracket)
                    if start_bracket != end_bracket:
                        line = next(in_file)
                #print(pool)
                for key, value in members.items():
                    if 'priority' not in value:
                        ws3.append([value, '=LEFT(A' + str(cell) + ',FIND(":",A' + str(cell) + ')-1)',
                                '=RIGHT(A' + str(cell) + ',LEN(A' + str(cell) + ')-FIND(":",A'
                                + str(cell) + '))'])
                    cell += 1
            line = next(in_file)
        if lbmethod == '':
            lbmethod = 'round-robin'
        pool['B'] = lbmethod
        pool.update(members)
        ws2.append(pool)

    if line.startswith('node ') and '*' not in line:
        nodename = line.split()[1].strip()
        node = {}
        node['A'] = nodename
        while line.startswith('}') is False:
            if 'monitor' in line and 'and' not in line:
                node['C'] = line.split()[-1].strip()
            elif 'monitor' in line and 'and' in line:
                node['C'] = line.split()[1].strip()
                node['D'] = line.split()[-1].strip()
            elif 'screen' in line:
                description = line.split()[-1].strip()
                node['B'] = description
            line = next(in_file)
        ws4.append(node)
    if line.startswith('monitor '):
        monitorlist = {}
        monitorlist['A'] = line.split()[1].strip()
        while line.startswith('}') is False:
            if 'defaults from' in line:
                monitorlist['B'] = line.split()[-1].strip()
            elif 'send ' in line:
                monitorlist['C'] = line.split('"')[1].strip()
            elif 'recv ' in line:
                monitorlist['D'] = line.split()[1].strip()
            line = next(in_file)
        ws5.append(monitorlist)
    if line.startswith('profile ') and 'persist' not in line:
        profile = {}
        profile['A'] = line.split()[2].strip()
        profile['B'] = line.split()[1].strip()
        ptype = line.split()[1].strip()
        prow = 3
        line = next(in_file)
        dfrom = line.split()[-1].strip()
        profile['C'] = line.split()[-1].strip()
        line = next(in_file)
        prflcol = 'D'
        while line.startswith('}') is False:
            if ptype == 'fastL4':
                if 'idle timeout' in line:
                    idl = line.split()[-1].strip
                    if idl != '300':
                        profile['D'] = line.rstrip().lstrip()
            elif ptype == 'ntlm':
                if 'key by ipaddr' in line:
                    keyip = line.strip()[-1].strip()
                    if keyip != 'disable':
                        profile['D'] = line.rstrip().lstrip()
            elif ptype == 'smtp':
                if 'security enabled' in line:
                    sec = line.split()[-1].strip()
                    if sec != 'disable':
                        profile['D'] = line.rstrip().lstrip()
            elif ptype == 'oneconnect':
                if 'source mask' in line:
                    srcmsk = line.split()[-1].strip()
                    if srcmsk != 'none':
                        profile['D'] = line.rstrip().lstrip()
            elif dfrom == 'http':
                if 'redirect rewrite' in line:
                    rdrw = line.split()[-1].strip()
                    if rdrw != 'none':
                        profile['D'] = line.rstrip().lstrip()
                if 'insert xforward' in line:
                    insfor = line.split()[-1].strip()
                    if insfor != 'disable':
                        profile['E'] = line.rstrip().lstrip()
            elif dfrom == 'http-wan-compression':
                if 'compress keep accept encoding' in line:
                    comp = line.split()[-1].strip()
                    if comp != 'disable':
                        profile['D'] = line.rstrip().lstrip()
            elif ptype == 'tcp' and dfrom != 'tcp-lan-optimized':
                if 'proxy options' in line:
                    pxyopt = line.split()[-1].strip()
                    if pxyopt != 'enable':
                        profile['D'] = line.rstrip().lstrip()
                if 'nagle' in line:
                    nagle = line.split()[-1].strip()
                    if nagle != 'disable':
                        profile['E'] = line.rstrip().lstrip()
                if 'ack on push' in line:
                    aop = line.split()[-1].strip()
                    if aop != 'enable':
                        profile['F'] = line.rstrip().lstrip()
                if 'proxy buffer low' in line:
                    pbl = line.split()[-1].strip()
                    if pbl != '32768':
                        profile['G'] = line.rstrip().lstrip()
                if 'proxy buffer high' in line:
                    pbh = line.split()[-1].strip()
                    if pbh != '49152':
                        profile['H'] = line.rstrip().lstrip()
                if 'idle timeout' in line:
                    it = line.split()[-1].strip()
                    if it != '300':
                        profile['I'] = line.rstrip().lstrip()
                if 'send buffer' in line:
                    sb = line.split()[-1].strip()
                    if sb != '65535':
                        profile['J'] = line.rstrip().lstrip()
                if 'recv window' in line:
                    rb = line.split()[-1].strip()
                    if rb != '65535':
                        profile['K'] = line.rstrip().lstrip()
            #print(prflcol)
            prow += 1
            line = next(in_file)
        pcolumn += 1
        ws6.append(profile)
    if line.startswith('profile ') and 'persist' in line:
        profile = {}
        profile['A'] = line.split()[2].strip()
        line = next(in_file)
        df = line.split()[-1].strip()
        if df == 'dest_addr' or df == 'hash' or df == 'source_addr' or df == 'cookie':
            profile['C'] = df
        else:
            profile['C'] =  df
        line = next(in_file)
        profile['B'] = line.replace('mode', '').rstrip().lstrip()
        prflcol = 'D'
        while line.startswith('}') is False:
            if 'timeout' in line:
                time = line.split()[-1].strip()
                if time != '180':
                    profile['D'] = line.strip()
            elif 'rule ' in line:
                profile['E'] = line.strip()
            elif 'mask' in line:
                mask = line.split()[-1].strip()
                if mask != 'none':
                    profile['E'] = line.strip()
            elif 'map proxies' in line or 'across services' in line:
                profile['F'] = line.strip()
            if prflcol != 'Z' and len(prflcol) == 1:
                prflcol = chr(ord(prflcol)+1)
            elif prflcol == 'Z':
                prflcol = 'AA'
            elif len(prflcol) == 2:
                prflcol = 'A'+ chr(ord(prflcol[1]) + 1)
            prow += 1
            line = next(in_file)
        prscolumn += 1
        ws7.append(profile)
    if line.startswith('snatpool '):
        snat = {}
        snat['A'] = line.split()[1].strip()
        line = next(in_file)
        col = 'B'
        while line.startswith('}') is False:
            if 'members' in line and '{' not in line:
                snat[col] = line.split()[-1].strip()
            elif 'members' in line and '{' in line:
                line = next(in_file)
            if '}' not in line and 'members' not in line:
                snat[col] = line.split()[0]
                col = chr(ord(col)+1)
            line = next(in_file)
        ws8.append(snat)
    if line.startswith('rule '):
        rule = {}
        rule['A'] = line.split()[1].strip()
        rule['C'] = []
        start_bracket, end_bracket = 1, 2
        while start_bracket != end_bracket:
            if start_bracket == 1 and end_bracket == 2:
                start_bracket = 0
                end_bracket = 0
            if 'rule' not in line:
                if 'matchclass' in line:
                    line = line.replace('matchclass', 'class match')
                    line = line.replace('$::', '/Common/')
                    rule['C'].append(line.strip())
                else:
                    rule['C'].append(line.strip())
            if '{' in line:
                for i in line:
                    if i == '{':
                        start_bracket += 1
            if '}' in line:
                for i in line:
                    if i == '}':
                        end_bracket += 1
            if start_bracket != end_bracket:
                line = next(in_file)
        rule['C'] = str(rule['C'])
        ws9.append(rule)
    if line.startswith('class '):
        dgrp = {}
        dgrp['A'] = line.split()[1].strip()
        dgrp['D'] = []
        while line.startswith('}') is False:
            if 'network' in line:
                dgrp['D'].append(line.split()[1].strip())
            elif 'host' in line:
                dgrp['D'].append(line.split()[1].strip() + '/32')
            line = next(in_file)
        dgrp['D'] = str(dgrp['D'])
        ws10.append(dgrp)

wb.save(dname)

from ciscoconfparse import CiscoConfParse
ef get_fex(worksheet, configfile):
    configfile.seek(0)
    worksheet.append(['Fex Number', 'Description', 'Pinning Max-Links', 'Hardware',
                      'QoS Policy', 'FCoE enabled', 'FC Ports'])

    parse = CiscoConfParse(f)
    all_fex = parse.find_objects('^fex')

    for child in all_fex:
        fexnumber = child.text.strip() 
        print('Parent: ', fexnumber)
        fex = {}
        fex['A'] = fexnumber
        for item in child.children:
            if 'hardware' in item.text:
                hardware = item.text.split()[1].strip()
                fex['D'] = hardware
                print('  Hardware: ', hardware)
            if 'policy' in item.text:
                qospolicy = item.text.split()[-1].strip()
                fex['E'] = qospolicy
                print('  QoS Policy: ', qospolicy)
            if 'pinning' in item.text:
                maxlinks = item.text.strip()
                fex['C'] = maxlinks
                print('  Pinning: ', maxlinks)
            if 'description' in item.text:
                description = item.text.split()[-1].strip('"')
                fex['B'] = description
                print('  Description: ', description)
        worksheet.append(fex)
